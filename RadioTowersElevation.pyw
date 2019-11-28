# pylint: disable=E0611
# pylint: disable=E1101
#pyi-makespec --onefile --icon=icon.ico --noconsole VEZRead.py
from PyQt5.QtWidgets import QApplication,QMainWindow,QWidget,QVBoxLayout,QHBoxLayout,QLabel,QListView,QPushButton,\
    QScrollArea,QSizePolicy, QTableWidgetItem,QSplitter, QFrame, QSizePolicy, QListView, QTableWidget, qApp, QAction,\
    QMessageBox,QFileDialog, QErrorMessage, QDoubleSpinBox, QSpacerItem, QLineEdit, QItemDelegate, QCheckBox,\
    QAbstractItemView, QShortcut, QDialog, QProgressDialog, QComboBox, QSpinBox, QCompleter, QRadioButton
from PyQt5.QtGui import QPixmap, QPalette, QBrush, QImage, QIcon, QTransform, QStandardItemModel,QStandardItem,\
     QDoubleValidator, QValidator, QCloseEvent, QColor, QIntValidator, QKeySequence
from PyQt5.QtCore import QPersistentModelIndex, Qt,  QSize, QModelIndex, QThread, pyqtSignal, QSortFilterProxyModel
import sys
import os

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
import matplotlib.ticker
import matplotlib
from matplotlib.ticker import FuncFormatter, FormatStrFormatter
import matplotlib.font_manager as fm
arial_font = fm.FontProperties(fname = "fonts/arial.ttf",family='Arial',
                                   weight='normal',
                                   style='normal', size=13)#

title_arial_font = fm.FontProperties(fname = "fonts/arial.ttf",family='Arial',
                                   weight='bold',
                                   style='normal', size=16)#

from functools import partial
import numpy as np

from time import time
from openpyxl import load_workbook, Workbook

import main

class CustomTable(QTableWidget):
    """ Добавляем сигнал завершения ввода  """
    def __init__(self,tb,parent=None):
        super().__init__(parent)
        self.Trig = False
        self.tb = tb
        self.editing = None
        self.edit_openfile = False

    def setcloseEditorSignal(self,metod):
        self.metod = metod
        self.Trig = True

    def closeEditor(self, editor, hint):
        if self.Trig:
            if self.tb == 1:
                self.metod()
            elif self.tb ==2:
                (lambda:self.metod(row=self.editing))()
        self.edit_openfile = True
        QTableWidget.closeEditor(self, editor, hint)

class MyValidator(QValidator):
    """ Позволяет вводить только числа """
    def __init__(self, var, parent,to=False,minus=False):
        QValidator.__init__(self, parent)
        self.minus = minus
        self.var = var
        self.to = to
        self.s = set(['0','1','2','3','4','5','6','7','8','9','.',',',''])

    def validate(self, s, pos): 
        """ проверяет привильная ли строка """   
        i=-1
        t1 = 0
        t2 = False
        t3 = 0
        for i in range(len(s)):
            if self.minus and i==0 and s[i] =="-":
                t3 += 1
                
            elif self.minus and i!=0 and s[i] =="-":
                i=-1
                break

            if self.to and i==2:
                if s[i] !="." and s[i] !="," and self.var == "duble":
                    i=-1
                    break
                elif self.var == "int":
                    i=-1
                    break
            if s[i] == ".":
                if self.var =="int":
                    i=-1
                    break
                elif self.var =="duble":
                    t1 += 1
            if s[i] == ",":
                if self.var =="int":
                    i=-1
                    break
                elif self.var =="duble":
                    t1 += 1
                    t2 = True
            if t1>1:
                i=-1
                break
            if s[i] not in self.s and not (self.minus and s[i]=="-"):
                i-=1
                break

        if s=='-':
            t2=True
        
        if i == len(s)-1:
            if t2:
                return (QValidator.Intermediate, s, pos) 
            else:
                return (QValidator.Acceptable, s, pos)
        else:
            return (QValidator.Invalid, s, pos)

    def fixup(self, s):
        """ форматирует неправильную строку """
        s1=''
        if s=="-":return ""
        t = False
        for i in s:
            if i in self.s or (self.minus and i=="-"):
                if  (i=="." or i==","):
                    if not t:
                        s1+="."
                        t = True
                else:
                    s1+=i
        s=s1
        return s

class DownloadDelegate(QItemDelegate):
    """ Переопределение поведения ячейки таблицы """
    def __init__(self, parent=None):
        super(DownloadDelegate, self).__init__(parent)

    def createEditor(self, parent, option, index):
        lineedit=QLineEdit(parent)
        if index.column() == 1 or index.column() == 4:
            lineedit.setValidator(MyValidator("duble",lineedit,minus=True))
        elif index.column() == 2 or index.column() == 5:
            lineedit.setValidator(MyValidator("int",lineedit,to=True))
        elif index.column() == 3 or index.column() == 6:
            lineedit.setValidator(MyValidator("duble",lineedit,to=True))
        elif index.column() == 7:
            lineedit.setValidator(MyValidator("duble",lineedit))
        elif 8<=index.column() <= 9:
            return
        #print(index.row())
        #print(index.column())
        return lineedit#QItemDelegate.createEditor(self, parent, option, index)

class DownloadDelegate2(QItemDelegate):
    """ Переопределение поведения ячейки таблицы """
    def __init__(self, parent=None):
        super(DownloadDelegate2, self).__init__(parent)

    def createEditor(self, parent, option, index):
        lineedit=QLineEdit(parent)
        if index.column() == 1:
            return lineedit
        else:
            return

class ExtendedComboBox(QComboBox):
    """ Переопределяем поведение QComboBox, чтобы можно было в его вписать свою строку """
    def __init__(self, parent=None):
        super(ExtendedComboBox, self).__init__(parent)

        self.setFocusPolicy(Qt.StrongFocus)
        self.setEditable(True)

        # add a filter model to filter matching items
        self.pFilterModel = QSortFilterProxyModel(self)
        self.pFilterModel.setFilterCaseSensitivity(Qt.CaseInsensitive)
        self.pFilterModel.setSourceModel(self.model())

        # add a completer, which uses the filter model
        self.completer = QCompleter(self.pFilterModel, self)
        # always show all (filtered) completions
        self.completer.setCompletionMode(QCompleter.UnfilteredPopupCompletion)
        self.setCompleter(self.completer)

        # connect signals
        self.lineEdit().textEdited[str].connect(self.pFilterModel.setFilterFixedString)
        self.completer.activated.connect(self.on_completer_activated)


    # on selection of an item from the completer, select the corresponding item from combobox 
    def on_completer_activated(self, text):
        if text:
            index = self.findText(text)
            self.setCurrentIndex(index)
            self.activated[str].emit(self.itemText(index))


    # on model change, update the models of the filter and completer as well 
    def setModel(self, model):
        super(ExtendedComboBox, self).setModel(model)
        self.pFilterModel.setSourceModel(model)
        self.completer.setModel(self.pFilterModel)


    # on model column change, update the model column of the filter and completer as well
    def setModelColumn(self, column):
        self.completer.setCompletionColumn(column)
        self.pFilterModel.setFilterKeyColumn(column)
        super(ExtendedComboBox, self).setModelColumn(column)

class SavePlots(QDialog):
    def __init__(self,curent,al,func,parent=None):
        super(SavePlots,self).__init__(parent)
        self.curent = curent
        self.al = al
        self.func = func
        self.setWindowTitle("Сохранить результаты")
        
        HspacerItem = [QSpacerItem(2, 2, QSizePolicy.Expanding, QSizePolicy.Minimum) for i in range(11)]
        VspacerItem = QSpacerItem(2, 2, QSizePolicy.Minimum, QSizePolicy.Expanding)

        lb1 = QLabel("Графики:")
        self.plots = QComboBox()
        self.plots.addItems(["Отдельные","Совместные"])
        BoxLayout1 = QHBoxLayout()
        BoxLayout1.addWidget(lb1)
        BoxLayout1.addItem(HspacerItem[0])
        BoxLayout1.addWidget(self.plots)

        lb2 = QLabel("Шрифт:")
        self.fonts = QComboBox()
        self.fonts.addItems(self.List_fonts())
        BoxLayout2 = QHBoxLayout()
        BoxLayout2.addWidget(lb2)
        BoxLayout2.addItem(HspacerItem[1])
        BoxLayout2.addWidget(self.fonts)
        
        lb3 = QLabel("Размер шрифта:")
        self.fontsize = QSpinBox()
        self.fontsize.setRange(1,99)
        self.fontsize.setSingleStep(1)
        self.fontsize.setValue(12)
        BoxLayout3 = QHBoxLayout()
        BoxLayout3.addWidget(lb3)
        BoxLayout3.addItem(HspacerItem[2])
        BoxLayout3.addWidget(self.fontsize)

        lb4 = QLabel("Путь:")
        self.ph = ExtendedComboBox()
        self.ph.setMinimumWidth(85)
        if self.curent==-1:
            self.ph.addItems(["Все"])
        else:
            self.ph.addItems(["Текущий","Все"])
        BoxLayout4 = QHBoxLayout()
        BoxLayout4.addWidget(lb4)
        BoxLayout4.addItem(HspacerItem[3])
        BoxLayout4.addWidget(self.ph)

        lb8 = QLabel("dpi:")
        self.dpi = QSpinBox()
        self.dpi.setRange(50,300)
        self.dpi.setSingleStep(10)
        self.dpi.setValue(300)
        self.dpi.setSuffix(" px")
        BoxLayout9 = QHBoxLayout()
        BoxLayout9.addWidget(lb8)
        BoxLayout9.addItem(HspacerItem[8])
        BoxLayout9.addWidget(self.dpi)

        lb9 = QLabel("Формат:")
        self.format = QComboBox()
        self.format.addItems(["jpg","png","pdf","svg","eps"])
        BoxLayout11 = QHBoxLayout()
        BoxLayout11.addWidget(lb9)
        BoxLayout11.addItem(HspacerItem[9])
        BoxLayout11.addWidget(self.format)

        self.rb1 = QRadioButton("Стандартный размер")
        self.rb1.setChecked(True)
        lb5 = QLabel("Лист:")
        self.A = QComboBox()
        self.A.addItems(["A6","A5","A4","A3","A2","A1","A0"])
        self.A.setCurrentText("A4")
        BoxLayout5 = QHBoxLayout()
        BoxLayout5.addWidget(lb5)
        BoxLayout5.addItem(HspacerItem[4])
        BoxLayout5.addWidget(self.A)

        self.rb2 = QRadioButton("Персональный размер")
        self.rb3 = QRadioButton("Массив данных в Excel")

        lb6 = QLabel("Шаг X на 1 см:")
        lb7 = QLabel("Шаг Y на 1 см:")

        self.X = QSpinBox()
        self.X.setRange(10,100000)
        self.X.setSingleStep(1)
        self.X.setValue(250)
        self.X.setSuffix(" м")
        BoxLayout6 = QHBoxLayout()
        BoxLayout6.addWidget(lb6)
        BoxLayout6.addItem(HspacerItem[5])
        BoxLayout6.addWidget(self.X)

        self.Y = QSpinBox()
        self.Y.setRange(1,100)
        self.Y.setSingleStep(1)
        self.Y.setValue(10)
        self.Y.setSuffix(" м")
        BoxLayout7 = QHBoxLayout()
        BoxLayout7.addWidget(lb7)
        BoxLayout7.addItem(HspacerItem[6])
        BoxLayout7.addWidget(self.Y)

        self.btn = QPushButton("Сохранить")
        self.btn.clicked.connect(self.Output)
        BoxLayout8 = QHBoxLayout()
        BoxLayout8.addItem(HspacerItem[7])
        BoxLayout8.addWidget(self.btn)

        BoxLayout10 = QVBoxLayout()
        BoxLayout10.addLayout(BoxLayout1)
        BoxLayout10.addLayout(BoxLayout2)
        BoxLayout10.addLayout(BoxLayout3)
        BoxLayout10.addLayout(BoxLayout4)
        BoxLayout10.addLayout(BoxLayout9)
        BoxLayout10.addLayout(BoxLayout11)
        BoxLayout10.addWidget(self.rb1)
        BoxLayout10.addLayout(BoxLayout5)
        BoxLayout10.addWidget(self.rb2)
        BoxLayout10.addLayout(BoxLayout6)
        BoxLayout10.addLayout(BoxLayout7)
        BoxLayout10.addWidget(self.rb3)
        BoxLayout10.addItem(VspacerItem)
        BoxLayout10.addLayout(BoxLayout8)

        self.setLayout(BoxLayout10)

        self.Asize = {"A6":(148,105),"A5":(210,148),"A4":(297,210),\
                "A3":(420,297),"A2":(594,420),"A1":(841,594),"A0":(1189,841)}
        self.sp=[]
        self.cl = False

    def List_fonts(self):
        files = os.listdir("fonts")
        l = [os.path.splitext(i)[0] for i in files if os.path.splitext(i)[1]==".ttf"]
        return l

    def Routes(self):
        text = self.ph.currentText()
        if text == "Текущий":
            return [self.curent]
        elif text == "Все":
            return [i for i in range(self.al)]
        else:
            sp=[]
            while True:
                i=text.find(',')
                if i!=-1:
                    sp.append(text[:i].replace(" ",''))
                    text = text[i+1:]
                else: break
            sp.append(text.replace(" ",''))
            routs = set()
            for s in sp:
                i=s.find('-')
                if i==-1:
                    if not s.isdigit(): raise Exception("Value outside range")
                    a = int(s)
                    if 1<=a<=self.al:
                        routs.add(a-1)
                    else: raise Exception("Value outside range")
                else:
                    if not s[:i].isdigit() or not s[i+1:].isdigit():
                        raise Exception("Value outside range")
                    a = int(s[:i])
                    b = int(s[i+1:])
                    if 1<=a<=self.al and 1<=b<=self.al and a<b:
                        for n in range(a-1,b):
                            routs.add(n) 
                    else: raise Exception("Value outside range")
            return list(routs)   

    def Output(self):    
        try: 
            self.sp.append(self.plots.currentText())
            self.sp.append("fonts/"+self.fonts.currentText()+".ttf")
            self.sp.append(self.fontsize.value())
            self.sp.append(self.Routes())
            self.sp.append(self.dpi.value())
            self.sp.append(self.format.currentText())
            if self.rb1.isChecked():
                self.sp.append(1)
                self.sp.append(self.Asize[self.A.currentText()])
            elif self.rb2.isChecked():
                self.sp.append(2)
                self.sp.append((self.X.value(),self.Y.value()))
            elif self.rb3.isChecked():
                self.sp.append(3)
                self.sp.append(self.Asize["A4"])

            self.cl = True
            self.close()
        except Exception as ex:
            if str(ex) == "Value outside range":
                ems = QErrorMessage(self)
                ems.setWindowTitle('Возникла ошибка')
                ems.showMessage('Некорректрно заданы значения в строке "Путь"')
            else:
                print(ex)

    def closeEvent(self, event):
        if self.cl:
            self.func(self.sp)
        




    
class Route(QDialog): #QWidget
    def __init__(self,name,row,func,parent=None):
        super(Route,self).__init__(parent)

        self.setWindowTitle('Путь - '+name)
        self.resize(600,300)
        self.setModal(True)
        self.row = row
        self.func = func

        self.leftlist = QListView()
        self.leftlist.setEditTriggers(QAbstractItemView.NoEditTriggers) # Запрещаем редактировнаие списка
        self.leftlist.pressed.connect(partial(self.Presseditem,1))
        self.left_mod = QStandardItemModel()
        self.d_left = {}
        self.id_left = {}
        #self.leftlist.setMovement(2) перемещение элементов мышкой

        self.rightlist = QListView()
        self.rightlist.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.rightlist.pressed.connect(partial(self.Presseditem,2))
        self.right_mod = QStandardItemModel()
        self.d_right = {}
        self.id_right = {}

        self.right = QPushButton("\u2192")
        self.right.clicked.connect(lambda:self.Left_Right(1))
        self.right_hot = QShortcut(QKeySequence("Ctrl+Right"), self)
        self.right_hot.activated.connect(lambda:self.Left_Right(1))
        self.left = QPushButton("\u2190")
        self.left.clicked.connect(lambda:self.Left_Right(-1))
        self.left_hot = QShortcut(QKeySequence("Ctrl+Left"), self)
        self.left_hot.activated.connect(lambda:self.Left_Right(-1))
        self.up = QPushButton("\u2191")
        self.up.clicked.connect(lambda:self.Down_Up(-1))
        self.up_hot = QShortcut(QKeySequence("Ctrl+Up"), self)
        self.up_hot.activated.connect(lambda:self.Down_Up(-1))
        self.down = QPushButton("\u2193")
        self.down.clicked.connect(lambda:self.Down_Up(1))
        self.down_hot = QShortcut(QKeySequence("Ctrl+Down"), self)
        self.down_hot.activated.connect(lambda:self.Down_Up(1))
        self.ok = QPushButton("Ok")
        self.ok.clicked.connect(self.Output)
        self.ok_hot = QShortcut(QKeySequence(Qt.Key_Return), self)
        self.ok_hot.activated.connect(self.Output)
        self.esc_hot = QShortcut(QKeySequence(Qt.Key_Escape), self)
        self.esc_hot.activated.connect(lambda:self.close())

        lb1 = QLabel("Маршрут:")
        lb2 = QLabel("Точки:")

        HspacerItem = [QSpacerItem(2, 2, QSizePolicy.Expanding, QSizePolicy.Minimum) for i in range(3)]
        VspacerItem = [QSpacerItem(2, 2, QSizePolicy.Minimum, QSizePolicy.Expanding) for i in range(2)]

        BoxLayout1 = QHBoxLayout()
        BoxLayout1.addItem(HspacerItem[0])
        BoxLayout1.addWidget(self.up)
        BoxLayout1.addWidget(self.down)
        BoxLayout1.addItem(HspacerItem[1])

        BoxLayout2 = QHBoxLayout()
        BoxLayout2.addItem(HspacerItem[2])
        BoxLayout2.addWidget(self.ok)

        BoxLayout3 = QVBoxLayout()
        BoxLayout3.addWidget(lb1)
        BoxLayout3.addWidget(self.leftlist)
        BoxLayout3.addLayout(BoxLayout1)

        BoxLayout4 = QVBoxLayout()
        BoxLayout4.addItem(VspacerItem[0])
        BoxLayout4.addWidget(self.left)
        BoxLayout4.addWidget(self.right)
        BoxLayout4.addItem(VspacerItem[1])

        BoxLayout5 = QVBoxLayout()
        BoxLayout5.addWidget(lb2)
        BoxLayout5.addWidget(self.rightlist)
        BoxLayout5.addLayout(BoxLayout2)

        BoxLayout6 = QHBoxLayout()
        BoxLayout6.addLayout(BoxLayout3,stretch=1)
        BoxLayout6.addLayout(BoxLayout4,stretch=0)
        BoxLayout6.addLayout(BoxLayout5,stretch=1)

        self.setLayout(BoxLayout6)

        self.adres = (0,-1,None)
        self.cl = False


    def Input(self,sp_left,sp_right):
        for i in range(len(sp_left)):
            a = QStandardItem(sp_left[i][1])
            self.left_mod.appendRow(a)
            Pm = QPersistentModelIndex(a.index())
            self.d_left[Pm]=a
            self.id_left[Pm] = sp_left[i][0]
        self.leftlist.setModel(self.left_mod)

        for i in range(len(sp_right)):
            a = QStandardItem(sp_right[i][1])
            self.right_mod.appendRow(a)
            Pm = QPersistentModelIndex(a.index())
            self.d_right[Pm]=a
            self.id_right[Pm] = sp_right[i][0]
        self.rightlist.setModel(self.right_mod)

        self.sp_left_old = sp_left
        self.sp_right_old = sp_right
        

    def Presseditem(self,lst, modelindex):
        modelindex=QPersistentModelIndex(modelindex)
        try:
            if lst == 1:
                self.adres = (lst,self.d_left[modelindex].row(),modelindex)
            elif lst == 2:
                self.adres = (lst,self.d_right[modelindex].row(),modelindex)
                #print("work",lst)
        except Exception as ex:
            print(ex)

    def Down_Up(self,b):
        if self.adres[0] == 1:
            l = len(self.d_left)
            if (l-1 > self.adres[1] and b==1) or (self.adres[1] > 0 and b==-1):
                del self.d_left[self.adres[2]]
                a = self.left_mod.takeRow(self.adres[1])[0]
                self.left_mod.insertRow(self.adres[1]+b,a)
                Pm = QPersistentModelIndex(a.index())
                self.d_left[Pm]=a
                self.id_left[Pm] = self.id_left.pop(self.adres[2])
                # Убираем выделение со старого элемета и присваеваем новому
                sel = self.leftlist.selectionModel()
                sel.select(a.index(),sel.Select)
                sel.select(self.left_mod.item(self.adres[1]).index(),sel.Deselect)

                self.adres = (1,self.adres[1]+b,Pm)


    def Left_Right(self,v):
        if self.adres[0] == 1 and v==1:
            del self.d_left[self.adres[2]]
            a = self.left_mod.takeRow(self.adres[1])[0]
            self.right_mod.insertRow(0,a)
            Pm = QPersistentModelIndex(a.index())
            self.d_right[Pm]=a
            self.id_right[Pm] = self.id_left.pop(self.adres[2])
            self.adres = (0,-1,None)

        elif self.adres[0] == 2  and v==-1:
            del self.d_right[self.adres[2]]
            a = self.right_mod.takeRow(self.adres[1])[0]
            self.left_mod.appendRow(a)
            Pm = QPersistentModelIndex(a.index())
            self.d_left[Pm]=a
            self.id_left[Pm] = self.id_right.pop(self.adres[2])
            self.adres = (0,-1,None)

        elif self.adres[0] == 0 and v==1:
            l = len(self.d_left)
            if l != 0:
                M = QPersistentModelIndex(self.left_mod.item(l-1).index())
                del self.d_left[M]
                a = self.left_mod.takeRow(l-1)[0]
                self.right_mod.insertRow(0,a)
                Pm = QPersistentModelIndex(a.index())
                self.d_right[Pm]=a
                self.id_right[Pm] = self.id_left.pop(M)
                self.adres = (0,-1,None)

        elif self.adres[0] == 0 and v==-1:
            if len(self.d_right) != 0:
                M = QPersistentModelIndex(self.right_mod.item(0).index())
                del self.d_right[M]
                a = self.right_mod.takeRow(0)[0]
                self.left_mod.appendRow(a)
                Pm = QPersistentModelIndex(a.index())
                self.d_left[Pm]=a
                self.id_left[Pm] = self.id_right.pop(M)
                self.adres = (0,-1,None)

    def Output(self):
        i = 0
        self.sp_left = []
        while self.left_mod.item(i):
            a = []
            item = self.left_mod.item(i)
            modelindex = QPersistentModelIndex(self.left_mod.indexFromItem(item))
            a.append(self.id_left[modelindex])
            a.append(item.text())
            self.sp_left.append(a)
            i+=1

        i = 0
        self.sp_right = []
        while self.right_mod.item(i):
            a = []
            item = self.right_mod.item(i)
            modelindex = QPersistentModelIndex(self.right_mod.indexFromItem(item))
            a.append(self.id_right[modelindex])
            a.append(item.text())
            self.sp_right.append(a)
            i+=1
        self.cl = True
        self.close()
    
    def closeEvent(self, event):
        if not self.cl:
            self.sp_left = self.sp_left_old
            self.sp_right = self.sp_right_old

        self.func(self.sp_left,self.sp_right,self.row)



class MyWindow(QMainWindow):
    def __init__(self,parent=None):
        super(MyWindow,self).__init__(parent)

        desktop = QApplication.desktop()
        wd = desktop.width()
        hg = desktop.height()
        ww = 1000
        wh = 500
        if ww>wd: ww = int(0.7*wd)
        if wh>hg: wh = int(0.7*hg)
        x = (wd-ww)//2
        y = (hg-wh)//2
        self.setGeometry(x, y, ww, wh)

        self.program_name = "Radio Towers Elevation"
        self.setWindowTitle(self.program_name)
        self.setWindowIcon(QIcon('icon.jpg'))

        try:
            self.path_home = os.path.expanduser("~\\Desktop\\")
        except Exception:
            self.path_home = ""

        for curren_dir in ["hgt"]:
            if os.path.exists(curren_dir):
                if os.path.isdir(curren_dir):
                    print(curren_dir+" is here")
                else:
                    try:
                        os.mkdir(curren_dir)
                    except OSError:
                        print ("Error generate dir "+curren_dir)
            else:
                try:
                    os.mkdir(curren_dir)
                except OSError:
                    print ("Error generate dir "+curren_dir)

        topVBoxLayout = QVBoxLayout(self) 
        #topVBoxLayout.setContentsMargins(0,0,0,0) 

        # Область для таблицы
        TbFrame = QFrame() 
        TbFrame.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding) 
        TbFrame.setMinimumSize(QSize(0, 200))

        self.Table = CustomTable(1)#QTableWidget()
        self.Table.setColumnCount(10)
        self.Table.setHorizontalHeaderLabels(["Название\nточки",\
                                                 "Широта\u00B0","'",'"',\
                                                "Долгота\u00B0","'",'"',\
                                                "Высотва\nвышки",\
                                                    "Высота над\nуровнем моря",\
                                                        "h_min"])
        
        self.Table.setColumnWidth(0,100)
        self.Table.setColumnWidth(1,70)
        self.Table.setColumnWidth(2,50)
        self.Table.setColumnWidth(3,60)
        self.Table.setColumnWidth(4,70)
        self.Table.setColumnWidth(5,50)
        self.Table.setColumnWidth(6,60)
        self.Table.setColumnWidth(7,70)
        self.Table.setColumnWidth(8,90)
        self.Table.setColumnWidth(9,55)
        #self.Table.cellChanged.connect(self.ClickCellChangedTable)
        self.Table.cellClicked[int,int].connect(self.ClickCellChangedTable)
        self.Table.setItemDelegate(DownloadDelegate(self))
        self.Table.setcloseEditorSignal(self.Calculate)
        BoxLayout1 = QHBoxLayout()
        BoxLayout1.addWidget(self.Table)
        TbFrame.setLayout(BoxLayout1)

        # Таблица путей
        RFrame = QFrame() 
        RFrame.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding) 
        RFrame.setMinimumSize(QSize(0, 10))

        self.TableV = CustomTable(2)#QTableWidget()
        #self.TableV.setcloseEditorSignal()
        self.TableV.setColumnCount(3)
        self.TableV.setItemDelegate(DownloadDelegate2(self))
        self.TableV.cellClicked[int,int].connect(self.ClickCellChangedTableV)
        self.TableV.setcloseEditorSignal(self.Calculate)
        self.TableV.setHorizontalHeaderLabels(["", "Название пути","Параметры пути"])
        self.TableV.setColumnWidth(0,25)
        self.TableV.setColumnWidth(1,100) 

        BoxLayout2 = QHBoxLayout()
        BoxLayout2.addWidget(self.TableV)
        RFrame.setLayout(BoxLayout2)

        # Область для настроек
        lb1 = QLabel("Высота запаса над уровнем земли")
        self.d_zp = QDoubleSpinBox()
        self.d_zp.setRange(0,999)
        self.d_zp.setSingleStep(0.1)
        self.d_zp.setSuffix(" м")
        self.d_zp.setDecimals(1)
        self.d_zp.setValue(20)
        self.d_zp.editingFinished.connect(lambda:self.Calculate(i=-1))

        lb2 = QLabel("Шаг получения значения высоты")
        self.d_step = QDoubleSpinBox()
        self.d_step.setRange(0,9999)
        self.d_step.setSingleStep(0.1)
        self.d_step.setSuffix(" м")
        self.d_step.setDecimals(1)
        self.d_step.setValue(30)
        self.d_step.editingFinished.connect(lambda:self.Calculate(i=-1))

        lb3 = QLabel("Частота сигнала")
        self.d_fg = QDoubleSpinBox()
        self.d_fg.setRange(1,99)
        self.d_fg.setSingleStep(0.1)
        self.d_fg.setSuffix(" ГГц")
        self.d_fg.setDecimals(1)
        self.d_fg.setValue(11)
        self.d_fg.editingFinished.connect(lambda:self.Calculate(i=-1))

        lb4 = QLabel("Зона Френеля")
        self.d_fgps = QSpinBox()
        self.d_fgps.setRange(0,100)
        self.d_fgps.setSingleStep(1)
        self.d_fgps.setSuffix(" %")
        self.d_fgps.setValue(100)
        self.d_fgps.editingFinished.connect(lambda:self.Calculate(i=-1))

        spacerItem = QSpacerItem(2, 2, QSizePolicy.Expanding, QSizePolicy.Minimum)
        #self.ch_h_min = QCheckBox("Расчитывать h_min по завершению ввода")

        BoxLayout3 = QHBoxLayout()
        BoxLayout3.addWidget(lb1)
        BoxLayout3.addWidget(self.d_zp)
        BoxLayout3.addWidget(lb2)
        BoxLayout3.addWidget(self.d_step)
        BoxLayout3.addWidget(lb3)
        BoxLayout3.addWidget(self.d_fg)
        BoxLayout3.addWidget(lb4)
        BoxLayout3.addWidget(self.d_fgps)
        BoxLayout3.addItem(spacerItem)
        
        
        Splitter1 = QSplitter(Qt.Horizontal) 
        Splitter1.addWidget(TbFrame)
        Splitter1.addWidget(RFrame)
        Splitter1.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        Splitter1.setStretchFactor(0, 5) # индекс элемента, фактор растяжения
        Splitter1.setStretchFactor(1, 2)

        # Область для графика
        PLFrame = QFrame() 
        PLFrame.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding) 
        PLFrame.setMinimumSize(QSize(0, 10))

        self.fig = plt.figure(dpi=75)
        self.FigCan = FigureCanvas(self.fig)
        BoxLayout5 = QVBoxLayout()
        BoxLayout5.addWidget(self.FigCan)
        PLFrame.setLayout(BoxLayout5)

        """ ax = self.fig.add_subplot(111) #
        ax.plot([1,2],[1,2]) """

        Splitter2 = QSplitter(Qt.Horizontal) 
        Splitter2.addWidget(PLFrame)
        Splitter2.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # Обьеденяем область графика с областями ввода данных
        Splitter3 = QSplitter(Qt.Vertical) 
        Splitter3.addWidget(Splitter1)
        Splitter3.addWidget(Splitter2)
        Splitter3.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        Splitter3.setStretchFactor(0, 1) # индекс элемента, фактор растяжения
        Splitter3.setStretchFactor(1, 2)

        topVBoxLayout.addLayout(BoxLayout3)
        topVBoxLayout.addWidget(Splitter3) 
        
        self.central_widget = QWidget()
        self.central_widget.setLayout(topVBoxLayout)
        self.setCentralWidget(self.central_widget) 

        self.statusBar()
        menubar = self.menuBar()

        NewFile = QAction("Новый файл", self)
        NewFile.setShortcut('Ctrl+Т')
        NewFile.setStatusTip('Создать новый файл')
        NewFile.triggered.connect(self.NewFile)

        OpenFile = QAction("Открыть файл", self)
        OpenFile.setShortcut('Ctrl+O')
        OpenFile.setStatusTip('Открыть файл исходных данных')
        OpenFile.triggered.connect(self.OpenFile)

        SaveFile = QAction("Сохранить файл", self)
        SaveFile.setShortcut('Ctrl+S')
        SaveFile.setStatusTip('Сохранить файл исходных данных')
        SaveFile.triggered.connect(self.SaveFile)

        NewPick = QAction("Добавить точку", self)
        NewPick.setShortcut('Ctrl+A')
        NewPick.setStatusTip('Добавить новую точку')
        NewPick.triggered.connect(self.NewPick)

        NewRoute = QAction("Добавить путь", self)
        NewRoute.setShortcut('Ctrl+Q')
        NewRoute.setStatusTip('Добавить новый путь')
        NewRoute.triggered.connect(self.NewRoute)

        DelPick = QAction("Удалить точку", self)
        DelPick.setShortcut('Ctrl+D')
        DelPick.setStatusTip('Удалить точку из списка')
        DelPick.triggered.connect(self.DelPick)

        DelRoute = QAction("Удалить путь", self)
        DelRoute.setShortcut('Ctrl+W')
        DelRoute.setStatusTip('Удалить путь из списка')
        DelRoute.triggered.connect(self.DelRoute)

        SPlot = QAction("Сохранить результаты", self)
        SPlot.setShortcut('Ctrl+1')
        SPlot.setStatusTip('Сохранить результаты расчётов')
        SPlot.triggered.connect(self.SPlot)

        fileMenu = menubar.addMenu('&Файл')
        fileMenu.addAction(NewFile)
        fileMenu.addAction(OpenFile)
        fileMenu.addAction(SaveFile)

        editMenu = menubar.addMenu('&Правка')
        editMenu.addAction(NewPick)
        editMenu.addAction(DelPick)
        editMenu.addAction(NewRoute)
        editMenu.addAction(DelRoute)

        resultMenu = menubar.addMenu('&Результаты')
        resultMenu.addAction(SPlot)

        self.t_ind = -1
        self.stack_table = []
        self.id_t = {}

        self.tv_ind = -1
        self.stack_table_v = []
        self.id_tv = {}

        self.route = [[],[],[]]
        self.id_route = {}

        self.kesh_calc = {}

        self.potokRasch = OneRasch(self.kesh_calc,self.d_step)
        self.potokRasch.mysignal.connect(self.End_calc, Qt.QueuedConnection)

        self.Progress = QProgressDialog('Производится расчёт','Отмена', 0, 0, parent=self)#
        self.Progress.setWindowTitle('Расчет')
        self.Progress.setMinimumDuration(0)
        self.Progress.setWindowModality(Qt.WindowModal)
        self.Progress.canceled.connect(self.ClosePotok)
        self.Progress.close()

        self.ad = False

        self.global_bloc = False


    def NewFile(self):
        if self.Table.edit_openfile or self.TableV.edit_openfile:
            Message = QMessageBox(QMessageBox.Question,  'Новый файл',
                "Сохранить предыдущие данные?", parent=self)
            Message.addButton('Да', QMessageBox.YesRole)
            Message.addButton('Нет', QMessageBox.NoRole)
            #Message.addButton('Сохранить', QMessageBox.ActionRole)
            reply = Message.exec()
            if reply == 0:
                self.SaveFile()
        
        self.Table.setRowCount(0)
        self.TableV.setRowCount(0)

        self.t_ind = -1
        self.stack_table = []
        self.id_t = {}

        self.tv_ind = -1
        self.stack_table_v = []
        self.id_tv = {}

        self.route = [[],[],[]]
        self.id_route = {}

        self.kesh_calc = {}

        self.Table.edit_openfile = False
        self.TableV.edit_openfile = False
        
        self.setWindowTitle(self.program_name+' - Новый файл')




    def OpenFile(self):
        """ Открываем файл исходных данных """
        try:
            if self.Table.edit_openfile or self.TableV.edit_openfile:
                Message = QMessageBox(QMessageBox.Question,  'Открыть файл',
                    "Сохранить предыдущие данные?", parent=self)
                Message.addButton('Да', QMessageBox.YesRole)
                Message.addButton('Нет', QMessageBox.NoRole)
                #Message.addButton('Сохранить', QMessageBox.ActionRole)
                reply = Message.exec()
                if reply == 0:
                    self.SaveFile()

            fname = QFileDialog.getOpenFileName(self, 'Открыть файл', self.path_home,'*.xlsx;*.xls')[0] # Обрати внимание на последний элемент
            name = os.path.splitext(os.path.split(fname)[1])[0]

            if fname == '': raise Exception("Не выбран файл")
            Katal = load_workbook(filename = fname)
            sheets = Katal.sheetnames
            ln = len(sheets)
            if ln==0: return
            Kat = Katal[sheets[0]]
            
            j=1
            while  Kat.cell(row=2,column=j).value!=None:
                j+=1

            zapas = Kat.cell(row=1,column=2).value
            step = Kat.cell(row=1,column=4).value
            fg = Kat.cell(row=1,column=6).value
            fgps = Kat.cell(row=1,column=7).value

            try:
                zapas = round(float(zapas),1)
            except Exception: None
            else:
                self.d_zp.setValue(zapas)

            try:
                step = round(float(step),1)
            except Exception: None
            else:
                self.d_step.setValue(step)

            try:
                fg = round(float(fg),1)
            except Exception: None
            else:
                self.d_fg.setValue(fg)

            try:
                fgps = int(fgps)
            except Exception: None
            else:
                self.d_fgps.setValue(fgps)

            i = 3
            sp_point = []
            while True:
                Trig = False
                for k in range(1,j):
                    if Kat.cell(row=i,column=k).value !=None:
                        Trig = True

                if Trig:
                    name_p =  Kat.cell(row=i,column=1).value

                    lat_g = Kat.cell(row=i,column=2).value
                    lat_m = Kat.cell(row=i,column=3).value
                    lat_s = Kat.cell(row=i,column=4).value

                    lon_g = Kat.cell(row=i,column=5).value
                    lon_m = Kat.cell(row=i,column=6).value
                    lon_s = Kat.cell(row=i,column=7).value

                    sp_point.append([str(name_p) if name_p!=None else "",\
                                    str(lat_g) if lat_g!=None else "",\
                                    str(int(lat_m)) if lat_m!=None else "",\
                                    str(lat_s) if lat_s!=None else "",
                                    str(lon_g) if lon_g!=None else "",\
                                    str(int(lon_m)) if lon_m!=None else "",\
                                    str(lon_s) if lon_s!=None else ""])
                    i+=1
                else:
                    break

            sp_route =[]
            sp_nam_route = []
            for k in range(8,j,2):
                a = {}
                sp_nam_route.append(str(Kat.cell(row=2,column=k).value))
                for n in range(3,i):
                    num = Kat.cell(row=n,column=k).value
                    h = Kat.cell(row=n,column=k+1).value
                    if "<class 'int'>" == str(type(num)):
                        a[num-1] = [n-3,str(h) if h!=None else ""]
                
                sp_route.append([a[it] for it in sorted(a.keys())])

            id_t = {i:i for i in range(len(sp_point))}
            id_tv = {i:i for i in range(len(sp_route))}

            route = [[],[],[]]
            id_route = {}
            row=-1
            
            for d in sp_route:
                row+=1
                ls1 = []
                ls2 = []
                ls3 = {}
                s_l = set()
                for i in range(len(d)):
                    ls1.append([d[i][0],sp_point[d[i][0]][0]])
                    ls3[d[i][0]]=[d[i][1],"",""]
                    s_l.add(d[i][0])
                    if d[i][0] not in id_route:
                        id_route[d[i][0]] = set()
                        id_route[d[i][0]].add((0,row,i))
                        id_route[d[i][0]].add((2,row,d[i][0]))
                    else:
                        id_route[d[i][0]].add((0,row,i))
                        id_route[d[i][0]].add((2,row,d[i][0]))
                route[0].append(ls1)
                route[2].append(ls3)

                j=-1
                for i in id_t:
                    if i not in s_l:
                        j+=1
                        ls2.append([i,sp_point[i][0]])
                        if i not in id_route:
                            id_route[i] = set()
                            id_route[i].add((1,row,j))
                        else:
                            id_route[i].add((1,row,j))

                route[1].append(ls2)


        except Exception as ex:
            if str(ex) != "Не выбран файл":
                ems = QErrorMessage(self)
                ems.setWindowTitle('Возникла ошибка')
                ems.showMessage('При открытий файла возникла ошибка. ('+str(ex)+')')
        else:
            self.t_ind=len(sp_point)-1
            self.Table.setRowCount(0)
            self.Table.setRowCount(self.t_ind+1)
            for j in range(self.t_ind+1):
                for i in range(0,7):
                    self.Table.setItem(j,i, QTableWidgetItem(sp_point[j][i]))
                self.Table.setItem(j,7, QTableWidgetItem(''))
                for i in range(8,10):
                    self.Table.setItem(j,i, QTableWidgetItem(''))
                    self.Table.item(j,i).setBackground(QColor(162,205,90))

            self.tv_ind=len(sp_route)-1
            self.TableV.setRowCount(0)
            self.TableV.setRowCount(self.tv_ind+1)

            for j in range(self.tv_ind+1):
                item = QTableWidgetItem()
                item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
                item.setCheckState(Qt.Unchecked)
                self.TableV.setItem(j, 0, item)

                self.TableV.setItem(j,1, QTableWidgetItem(sp_nam_route[j]))
                self.TableV.setItem(j,2, QTableWidgetItem('Выбрать путь'))
                self.TableV.item(j,2).setBackground(QColor("#E1E1E1"))
            
            self.id_t = id_t 
            self.id_tv = id_tv
            self.stack_table = []
            self.stack_table_v = []

            self.route = route
            self.id_route = id_route

            self.kesh_calc = {}
            self.Table.edit_openfile = False
            self.TableV.edit_openfile = False

            self.setWindowTitle(self.program_name+' - '+name)

    def SaveFile(self):
        wb = Workbook()
        ws=wb.active
        try:
            fname = QFileDialog.getSaveFileName(self, 'Сохранить файл', self.path_home,'*.xlsx;;*.xls')[0]
            self.Zapis()
            self.ZapisV()
            ws.append(["Высота запаса, м",self.d_zp.value(),
                        "Шаг, м", self.d_step.value(),"Частота сигнала, ГГц",self.d_fg.value(),
                        "Зона Френеля, %",self.d_fgps.value()])
            ws.append(["Название точки",'Широта градусы','Минуты','Секунды','Долгота градусы','Минуты','Секунды'])
            for i in range(self.t_ind+1):
                ws.append([self.t_d[i][0],\
                            float(self.t_d[i][1]) if self.t_d[i][1]!='' else '',
                            int(self.t_d[i][2]) if self.t_d[i][2]!='' else '',
                            float(self.t_d[i][3]) if self.t_d[i][3]!='' else '',
                            float(self.t_d[i][4]) if self.t_d[i][4]!='' else '',
                            int(self.t_d[i][5]) if self.t_d[i][5]!='' else '',
                            float(self.t_d[i][6]) if self.t_d[i][6]!='' else ''])
            #ws.cell(row=1,column=1).value = "Название точки"
            d = {v:k for k,v in self.id_t.items()}
            k = 8
            for i in range(len(self.route[0])):
                ws.cell(row=2,column=k).value = self.t_v[i][1]
                ws.cell(row=2,column=k+1).value = 'Высота, м'
                for j in range(len(self.route[0][i])):
                    a=d[self.route[0][i][j][0]]
                    ws.cell(row=a+3,column=k).value = j+1
                    ws.cell(row=a+3,column=k+1).value = float(self.route[2][i][a][0]) if self.route[2][i][a][0]!='' else ''
                k+=2

            wb.save(fname)
            name = os.path.splitext(os.path.split(fname)[1])[0]
            self.setWindowTitle(self.program_name+' - '+name)
        except Exception as ex:
            ems = QErrorMessage(self)
            ems.setWindowTitle('Возникла ошибка')
            ems.showMessage('Возможно открыт файл в который осуществлиется попытка сохранения. Закройте файл. ('+str(ex)+')')
        else:
            QMessageBox.information(self, 'Сохранение','Операция прошла успешно.',
                                          buttons=QMessageBox.Ok,
                                          defaultButton=QMessageBox.Ok)
        


    def NewPick(self):
        """ Добавляем новую строку в таблицу точек """
        ind=self.Videl_str(self.Table.selectedIndexes())
        self.Zapis()
        self.t_ind+=1
        self.Table.setRowCount(self.t_ind+1)

        
        for i in range(0,8):
            self.Table.setItem(self.t_ind,i, QTableWidgetItem(''))

        for i in range(8,10):
            self.Table.setItem(self.t_ind,i, QTableWidgetItem(''))
            self.Table.item(self.t_ind,i).setBackground(QColor(162,205,90))

        
        if ind != -1 and ind != self.t_ind:
            sp_id = []
            for i in range(ind+1,self.t_ind+1):
                sp_id.append([i,self.id_t.pop(i-1)])      
                for j in range(0,10):
                    self.Table.item(i, j).setText(self.t_d[i-1][j])

            self.id_t.update(sp_id)
            self.id_t[ind] = self.t_ind if len(self.stack_table)==0 else self.stack_table.pop()
            for r in self.route[2]:
                if r != None:
                    r[self.id_t[ind]] = ['','','']

            for j in range(0,10):
                self.Table.item(ind, j).setText("")

        else:
            self.id_t[self.t_ind] = self.t_ind if len(self.stack_table)==0 else self.stack_table.pop()
            for r in self.route[2]:
                if r != None:
                    r[self.id_t[self.t_ind]] = ['','','']

        #print(self.id_t)



    def NewRoute(self):
        """ Добавляем новую строку в таблицу Маршрутов """
        self.ad = True
        ind=self.Videl_str(self.TableV.selectedIndexes())
        self.ZapisV()
        self.tv_ind+=1
        self.TableV.setRowCount(self.tv_ind+1)

        item = QTableWidgetItem()
        item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
        item.setCheckState(Qt.Unchecked)
        self.TableV.setItem(self.tv_ind, 0, item)

        self.TableV.setItem(self.tv_ind,1, QTableWidgetItem(''))
        self.TableV.setItem(self.tv_ind,2, QTableWidgetItem('Выбрать путь'))
        self.TableV.item(self.tv_ind,2).setBackground(QColor("#E1E1E1"))

        if ind != -1 and ind != self.tv_ind:
            self.route[0].insert(ind,None)
            self.route[1].insert(ind,None)
            self.route[2].insert(ind,None)
            sp_id = []
            for i in range(ind+1,self.tv_ind+1): 
                sp_id.append([i,self.id_tv.pop(i-1)])
                item = QTableWidgetItem()
                item.setCheckState(Qt.Unchecked)
                item.setCheckState(Qt.Checked if self.t_v[i-1][0] else Qt.Unchecked) 
                self.TableV.setItem(i, 0, item)         
                self.TableV.item(i, 1).setText(self.t_v[i-1][1])
                self.TableV.item(i, 2).setText('Выбрать путь')

            self.id_tv.update(sp_id)
            self.id_tv[ind] = self.tv_ind if len(self.stack_table_v)==0 else self.stack_table_v.pop()

            item = self.TableV.item(ind, 0)
            item.setCheckState(Qt.Unchecked)
            self.TableV.item(ind, 1).setText("")
            self.TableV.item(ind, 2).setText('Выбрать путь')

        else:
            self.id_tv[self.tv_ind] = self.tv_ind if len(self.stack_table_v)==0 else self.stack_table_v.pop()
            self.route[0].append(None)
            self.route[1].append(None)
            self.route[2].append(None)

        self.ad = False
                
    def DelPick(self):
        """ Удаляем выделенную сроку таблицы точек"""
        if self.t_ind == -1: 1
        else:
            ind=self.Videl_str(self.Table.selectedIndexes()) # Получаем индекс выделенную строку

            st = ind if ind != -1 and ind != self.t_ind else self.t_ind

            Message = QMessageBox(QMessageBox.Question,  'Удаление',
                    "Вы дейстивлеьно хотите удалить " +str(st+1)+ " строку таблицы точек?", parent=self)
            Message.addButton('Да', QMessageBox.YesRole)
            Message.addButton('Нет', QMessageBox.NoRole)
            reply = Message.exec()       
            if reply == 0:

                if ind != -1 and ind != self.t_ind:
                    self.Zapis()
                    self.t_ind-=1
                    self.Table.setRowCount(self.t_ind+1)

                    idd = self.id_t.pop(ind)
                    self.stack_table.append(idd)
                    self.Del_id_point(idd)

                    sp_id=[]
                    for i in range(ind,self.t_ind+1):
                        sp_id.append([i,self.id_t.pop(i+1)])  
                        for j in range(0,10):
                            self.Table.item(i, j).setText(self.t_d[i+1][j])

                    self.id_t.update(sp_id)


                else:    
                    self.Del_id_point(self.id_t.pop(self.t_ind))

                    self.t_ind-=1
                    self.Table.setRowCount(self.t_ind+1)

                #self.ReadTable()

    def DelRoute(self):
        """ Удаляем выделеную строку таблицы путей """   
        if self.tv_ind == -1: 1
        else:
            ind=self.Videl_str(self.TableV.selectedIndexes()) # Получаем индекс выделенную строку

            st = ind if ind != -1 and ind != self.tv_ind else self.tv_ind

            Message = QMessageBox(QMessageBox.Question,  'Удаление',
                    "Вы дейстивлеьно хотите удалить " +str(st+1)+ " строку таблицы путей?", parent=self)
            Message.addButton('Да', QMessageBox.YesRole)
            Message.addButton('Нет', QMessageBox.NoRole)
            reply = Message.exec()       
            if reply == 0:
                if ind != -1 and ind != self.tv_ind:
                    self.ZapisV()
                    self.tv_ind-=1
                    self.TableV.setRowCount(self.tv_ind+1)

                    self.stack_table_v.append(self.id_tv.pop(ind))
                    self.Del_id_route(ind)
                    sp_id = []
                    for i in range(ind,self.tv_ind+1):
                        sp_id.append([i,self.id_tv.pop(i+1)]) 
                        item = QTableWidgetItem()
                        item.setCheckState(Qt.Unchecked)
                        item.setCheckState(Qt.Checked if self.t_v[i+1][0] else Qt.Unchecked) 
                        self.TableV.setItem(i, 0, item)         
                        self.TableV.item(i, 1).setText(self.t_v[i+1][1])
                        self.TableV.item(i, 2).setText('Выбрать путь')

                    self.id_tv.update(sp_id)
                else:    
                    del self.id_tv[self.tv_ind]

                    self.tv_ind-=1
                    self.TableV.setRowCount(self.tv_ind+1) 

                    self.Del_id_route(self.tv_ind+1)


    def Zapis(self):
        """ Записываем данные введённые в таблицу """
        self.t_d = []
        for i in range(self.t_ind+1):
            a = []
            for j in range(0,10):
                a.append(self.Table.item(i, j).text())
            self.t_d.append(a)
        #print(self.t_d)

    def ZapisV(self):
        """ Записываем данные введённые в таблицу """
        self.t_v = []
        for i in range(self.tv_ind+1):
            a = []
            item = self.TableV.item(i, 0)
            currentState = item.checkState()
            if currentState == Qt.Checked:
                a.append(True)
            else:
                a.append(False)
            for j in range(1,2):
                a.append(self.TableV.item(i, j).text())
            self.t_v.append(a)


    def Videl_str(self, sp):
        """ Получаем индекс выделенной строки """
        if len(sp) != 0:
            ind=sp[0].row()
            for i in range(1,len(sp)):
                if ind != sp[i].row():
                    ind=-1
                    break
        else:
            ind=-1
        return ind

    
    def ClickCellChangedTable(self, row, column):
        """ События нажатия на ячейки таблицы точек """
        1
        """ Проверяем состояние чекбоксов """
        """ if column == 0:
            self.ReadTable() """
        """ item = self.Table.item(row, column)
        currentState = item.checkState()
        if currentState == Qt.Checked:
            print("checked")
        else:
            print("unchecked") """

    def ClickCellChangedTableV(self, row, column):
        """ События нажатия на ячейки таблицы путей """
        if column == 2 and not self.ad:
            try:
                self.ZapisV()
                a=Route(self.t_v[row][1],row,lambda a,b,c:self.Read_Route(a,b,c),parent=self)
                d_names = self.List_Point()
                if self.route[0][row] == None or self.route[1][row] == None:
                    a.Input([],list(d_names.items()))
                    a.show()
                else:
                    for i in range(len(self.route[0][row])):
                        self.route[0][row][i][1] = d_names.pop(self.route[0][row][i][0])
                    for i in range(len(self.route[1][row])):
                        self.route[1][row][i][1] = d_names.pop(self.route[1][row][i][0])
                    
                    a.Input(self.route[0][row],self.route[1][row]+list(d_names.items()))
                    a.show()
            except Exception as ex:
                print(ex)

        elif column == 1 and not self.ad:
            self.TableV.editing = row

        elif column == 0 and not self.ad:
            item = self.TableV.item(row, column)
            currentState = item.checkState()
            if currentState == Qt.Checked:
                for i in range(self.tv_ind+1):
                    if i==row: continue
                    self.TableV.item(i, 0).setCheckState(Qt.Unchecked)

                self.Calculate(row,call=True)
                


    def List_Point(self):
        """ Словарь id:имя точки и обновление имён точек"""
        self.Zapis()
        d ={}
        for i in range(self.t_ind+1):
            d[self.id_t[i]]=self.t_d[i][0]
        return d

    def Read_Route(self,sp_left,sp_right,row):
        """ Обновляем параметры маршрута после его редактирования """
        Trig = False
        if self.route[2][row] == None:
            Trig = True
        else:
            D = self.route[2][row]
            
        self.Del_id_route(row)

        self.route[0].insert(row,sp_left)
        self.route[1].insert(row,sp_right)
        self.route[2].insert(row,{})
        #print(self.route[1][row])

        d = {v:k for k,v in self.id_t.items()}

        for i in range(len(sp_left)):
            if Trig:
                self.route[2][row][sp_left[i][0]]=[self.t_d[d[sp_left[i][0]]][7],\
                    self.t_d[d[sp_left[i][0]]][8],self.t_d[d[sp_left[i][0]]][9]]
            else:
                self.route[2][row][sp_left[i][0]] = D.get(sp_left[i][0],["","",""])

            if sp_left[i][0] not in self.id_route:
                self.id_route[sp_left[i][0]] = set()
                self.id_route[sp_left[i][0]].add((0,row,i))
                self.id_route[sp_left[i][0]].add((2,row,sp_left[i][0]))

            else:
                self.id_route[sp_left[i][0]].add((0,row,i))
                self.id_route[sp_left[i][0]].add((2,row,sp_left[i][0]))
        

        for i in range(len(sp_right)):
            if sp_right[i][0] not in self.id_route:
                self.id_route[sp_right[i][0]] = set()
                self.id_route[sp_right[i][0]].add((1,row,i))

            else:
                self.id_route[sp_right[i][0]].add((1,row,i))               


        if self.t_v[row][0]:
            self.Calculate()

    
    def Del_id_route(self,row):
        """ Удаляем данные по маршруту поле удаления его строки в таблице """

        sp_left = self.route[0].pop(row)
        sp_right = self.route[1].pop(row)

        del self.route[2][row]

        if sp_left!= None:
            for i in range(len(sp_left)):
                if sp_left[i][0]  in self.id_route:
                    self.id_route[sp_left[i][0]].discard((0,row,i))
                    self.id_route[sp_left[i][0]].discard((2,row,sp_left[i][0]))

        if sp_right!= None:
            for i in range(len(sp_right)):
                if sp_right[i][0]  in self.id_route:
                    self.id_route[sp_right[i][0]].discard((1,row,i))



    def Del_id_point(self,ind):
        """ Удаляем упоминания об точке во всех маршрутах """
        d =self.id_route.pop(ind,None)
        if d!=None:
            d = list(d)
            for i in d:
                self.route[i[0]][i[1]].pop(i[2])


    def ReadTable(self,i,idd,H=None):
        try:
            a = []
            a.append(idd)
            a.append(self.t_d[i][0])
            if self.t_d[i][1]!='' and self.t_d[i][2]!='' and self.t_d[i][3]!='':
                mn = int(self.t_d[i][2])
                sc = float(self.t_d[i][3])
                if mn>60 or sc>60:return
                lat = main.DMS_in_DD(int(self.t_d[i][1]),mn,sc)
            elif self.t_d[i][1]!='' and self.t_d[i][2]=='' and self.t_d[i][3]=='':
                lat = float(self.t_d[i][1])
            else:
                print("Неработаем")
                return
            if lat>180:return
            if self.t_d[i][4]!='' and self.t_d[i][5]!='' and self.t_d[i][6]!='':
                mn = int(self.t_d[i][5])
                sc = float(self.t_d[i][6])
                if mn>60 or sc>60:return
                lon = main.DMS_in_DD(int(self.t_d[i][4]),mn,sc)
            elif self.t_d[i][4]!='' and self.t_d[i][5]=='' and self.t_d[i][6]=='':
                lon = float(self.t_d[i][4])
            else:
                print("Неработаем")
                return
            if lon>180:return
            a.append((lat,lon))
            if H != None:
                a.append(float(H))
            else:
                if self.t_d[i][7]!='':
                    a.append(float(self.t_d[i][7]))
                else:
                    print("Неработаем")
                    return 
            return a
        except Exception:
            return

    def Write_h(self,D_h):
        """ Записываем инфу об высотах пути в таблицу точек """
        if D_h != None:
            for i in range(self.t_ind+1):
                if self.id_t[i] in D_h:
                    sp = D_h[self.id_t[i]]
                    self.Table.item(i,7).setText(sp[0])
                    self.Table.item(i,7).setBackground(QColor(46,117,182))
                    self.Table.item(i,8).setText(sp[1])
                    self.Table.item(i,9).setText(sp[2])
                else:
                    self.Table.item(i,7).setText("")
                    self.Table.item(i,7).setBackground(QColor(255,255,255))
                    self.Table.item(i,8).setText("")
                    self.Table.item(i,9).setText("")

                    #self.TableV.item(i, 2).setText('Выбрать путь')


    def Calculate(self,i=-1,call=False,row=-1,prnt=False):
        if self.ad: return
        self.ZapisV()
        Trig = False
        if i ==-1:
            for i in range(len(self.t_v)):
                if self.t_v[i][0]:
                    Trig = True
                    break
        else:
            Trig = True
        if row!=-1 and row!=i: return

        if Trig and self.route[0][i]!=None:
            if call and not prnt:
                self.Write_h(self.route[2][i])

            self.Zapis()
            d = {v:k for k,v in self.id_t.items()}

            sp_route = []
            if not call: self.route[2][i] = {}

            for j in range(len(self.route[0][i])):
                if not call: self.route[2][i][self.route[0][i][j][0]] = self.t_d[d[self.route[0][i][j][0]]][7:10]
                if prnt: val = self.ReadTable(d[self.route[0][i][j][0]],self.route[0][i][j][0],H=self.route[2][i][self.route[0][i][j][0]][0])
                else: val = self.ReadTable(d[self.route[0][i][j][0]],self.route[0][i][j][0])
                if val != None:
                    sp_route.append(val)

            if not prnt:
                if not self.potokRasch.isRunning() and not self.global_bloc:
                    self.global_bloc = True
                    self.potokRasch.InputDate([sp_route],d,[i],False)
                    print("Запускаем расчёт")
                    self.makeplot = True
                    self.potokRasch.start()
                    self.Progress.show()
            else:
                return sp_route





    def Relef(self,sp,Xrd,h1,h2,zapas,title='',t1=None,t2=None):
        #self.fig.add_subplot(111).clear()
        ax = self.fig.add_subplot(111) #
        #print(Xrd)
        #gr = [sp[0][1]*1.1,sp[12][1]*1.1,0,max(sp[1][1],sp[15][1],np.max(sp[9]))*1.1]

        line2, = ax.plot(sp[2][0]+Xrd, sp[2][1],"g")
        ax.plot(sp[2][2]+Xrd, sp[2][3],"g")

        h1_xy = sp[3][0](h1)
        h2_xy = sp[3][1](h2)

        lx_f1 ,ly_f1, lx_f2 ,ly_f2 = main.Frenel(sp[2][6]*1000,h1_xy,h2_xy,self.d_fg.value(),self.d_fgps.value())

        ax.plot(lx_f1+Xrd, ly_f1,"r")
        ax.plot(lx_f2+Xrd, ly_f2,"r")

        line1, = ax.plot([sp[2][0][1]+Xrd,h1_xy[0]+Xrd], [sp[2][1][1],h1_xy[1]],"b")
        ax.plot([sp[2][2][1]+Xrd,h2_xy[0]+Xrd], [sp[2][3][1],h2_xy[1]],"b")

        line5, = ax.plot([h1_xy[0]+Xrd,h2_xy[0]+Xrd], [h1_xy[1],h2_xy[1]],"r--")

        line3, = ax.plot(sp[2][4]+Xrd, sp[2][5],"k")
        #print(max(sp[2][5]))
        ax.plot(sp[2][6]+Xrd, sp[2][7],"g")
        line4, = ax.plot(sp[2][6]+Xrd, sp[2][7]+zapas,"y")

        hr = np.max(sp[2][5])

        text = 13
        tit = 16
        #leg = 14
        tiks = 13
        label_xy = 13

        ax.text(Xrd,np.max(sp[2][5]),str(round(sp[1],2))+" км\n"+str(round(hr,2))+" м",\
            horizontalalignment='center', verticalalignment='bottom',fontproperties=arial_font,fontsize=text)

        if t1!=None:
            ax.text(sp[2][0][0]+Xrd,sp[6][0]/2,t1,horizontalalignment='right',\
                verticalalignment='center',rotation='vertical',fontproperties=arial_font,fontsize=text)
            ax.text(h1_xy[0]+Xrd,(h1_xy[1]-sp[2][1][1])/2+sp[2][1][1],str(round(h1,2))+" м",\
                horizontalalignment='left',verticalalignment='center',rotation='vertical',
                fontproperties=arial_font,fontsize=text)
        if t2!=None:
            ax.text(sp[2][2][0]+Xrd,sp[6][1]/2,t2,horizontalalignment='right',\
                verticalalignment='center',rotation='vertical',fontproperties=arial_font,fontsize=text)
            ax.text(h2_xy[0]+Xrd,(h2_xy[1]-sp[2][3][1])/2+sp[2][3][1],str(round(h2,2))+" м",\
                horizontalalignment='left',verticalalignment='center',rotation='vertical',\
                    fontproperties=arial_font,fontsize=text)
            ax.set_title(title,fontproperties=title_arial_font,fontsize=tit)

            self.fig.legend(handles=(line1, line2, line3,line4,line5),\
                        labels=('Вышки', 'Высота\nн.у.м.', 'Rземли','+'+str(zapas)+' м','Сигнал'),
                        loc='right',prop=arial_font)#

        ax.grid(True) # Включаем сетку
        ax.set_xlabel('км',fontproperties=arial_font,fontsize=label_xy) # Подпись оси х
        ax.set_ylabel('м',fontproperties=arial_font,fontsize=label_xy) # Подпись оси у

        for label in (ax.get_xticklabels() + ax.get_yticklabels()):
            label.set_fontproperties(arial_font)
            label.set_fontsize(tiks) # Size here overrides font_prop

        
        #ax.axis(gr)
        self.FigCan.draw()
        #ax.axis([-25120-150,-25120+150,0,300])
        #3508х4961
        #s = ''
        #ax.set_yticks(np.arange(0,gr[3],10))
        #ax.set_xticks(np.hstack((np.flip(np.arange(0,gr[0],-500)),np.arange(gr[0],gr[1],500))))
        
        #Deistv.savefig("Графики/"+s+'.png', format='png', dpi=400,figsize=(12,8))
        #Deistv.show()
        
              
    def ClosePotok(self):
        self.potokRasch.stop()
        #self.potokRasch.mysignal.disconnect()
        #del self.potokRasch
        
    def End_calc(self,mess):
        if self.makeplot:
           self.WinPlot(mess)
        else:
            self.PrintPlot(mess)

#PointElevSaveExcel(self,wb,title,t1,t2,sp,ws,newpage=False):            

    def PrintPlot(self,mess):
        sp_list, sp_route, d, i = self.potokRasch.OutputDate()
        self.Progress.close()
        if self.print_param[6]==3:
            wb = Workbook()
            ws=None
        else:
            fig = plt.figure() #dpi=75,figsize=(8,6)
        try:
            if mess!="End": raise Exception(mess)
            for n in range(len(sp_list)):
                if sp_route[n]==[]: continue
                Xrd = 0
                if self.print_param[6]!=3:fig.clear()
                ln = len(sp_route[n])
                j=0
                for sp in sp_list[n]:
                    j+=1
                    if self.print_param[6]!=3:
                        if self.print_param[0]=="Совместные":
                            Xrd += sp[0]/2
                            if j == ln-1:
                                self.PrintRelef(fig,sp,Xrd,sp_route[n][j-1][3],\
                                    sp_route[n][j][3],self.d_zp.value(),title=self.t_v[i[n]][1],t1=sp_route[n][j-1][1],t2=sp_route[n][j][1])
                            else:
                                self.PrintRelef(fig,sp,Xrd,sp_route[n][j-1][3],\
                                    sp_route[n][j][3],self.d_zp.value(),t1=sp_route[n][j-1][1])
                            Xrd += sp[0]/2
                        else:
                            fig.clear()
                            Xrd = sp[0]/2
                            self.PrintRelef(fig,sp,Xrd,sp_route[n][j-1][3],\
                                    sp_route[n][j][3],self.d_zp.value(),title=self.t_v[i[n]][1],t1=sp_route[n][j-1][1],t2=sp_route[n][j][1])

                    else:
                        if j==1:
                            ws=self.PointElevSaveExcel(wb,self.t_v[i[n]][1],sp_route[n][j-1][1],sp_route[n][j][1],sp,ws,newpage=True)
                        else:
                            ws=self.PointElevSaveExcel(wb,self.t_v[i[n]][1],sp_route[n][j-1][1],sp_route[n][j][1],sp,ws)
            if self.print_param[6]==3:wb.save(self.print_param[8])
            self.global_bloc=False
        except Exception as ex:
            self.global_bloc=False
            ems = QErrorMessage(self)
            ems.setWindowTitle('Возникла ошибка')
            ems.showMessage('Возникла следующая ошибка: ('+str(ex)+')')
        else:
            QMessageBox.information(self, 'Сохранение','Операция прошла успешно.',
                                          buttons=QMessageBox.Ok,
                                          defaultButton=QMessageBox.Ok)

    def PrintRelef(self,fig,sp,Xrd,h1,h2,zapas,title='',t1=None,t2=None):
        ax = fig.add_subplot(111) #
        font = fm.FontProperties(fname = self.print_param[1],
                                   weight='normal',
                                   style='normal', size=self.print_param[2])#


        line2, = ax.plot(sp[2][0]+Xrd, sp[2][1],"g")
        ax.plot(sp[2][2]+Xrd, sp[2][3],"g")

        h1_xy = sp[3][0](h1)
        h2_xy = sp[3][1](h2)

        lx_f1 ,ly_f1, lx_f2 ,ly_f2 = main.Frenel(sp[2][6]*1000,h1_xy,h2_xy,self.d_fg.value(),self.d_fgps.value())

        ax.plot(lx_f1+Xrd, ly_f1,"r")
        ax.plot(lx_f2+Xrd, ly_f2,"r")

        line1, = ax.plot([sp[2][0][1]+Xrd,h1_xy[0]+Xrd], [sp[2][1][1],h1_xy[1]],"b")
        ax.plot([sp[2][2][1]+Xrd,h2_xy[0]+Xrd], [sp[2][3][1],h2_xy[1]],"b")

        line5, = ax.plot([h1_xy[0]+Xrd,h2_xy[0]+Xrd], [h1_xy[1],h2_xy[1]],"r--")

        line3, = ax.plot(sp[2][4]+Xrd, sp[2][5],"k")

        ax.plot(sp[2][6]+Xrd, sp[2][7],"g")
        line4, = ax.plot(sp[2][6]+Xrd, sp[2][7]+zapas,"y")

        text = self.print_param[2]
        tit = self.print_param[2]+2
        tiks = self.print_param[2]
        label_xy = self.print_param[2]
        #leg = 14

        hr = np.max(sp[2][5])

        ax.text(Xrd,np.max(sp[2][5]),str(round(sp[1],2))+" км\n"+str(round(hr,2))+" м",\
            horizontalalignment='center', verticalalignment='bottom',fontproperties=arial_font,fontsize=text)

        if t1!=None:
            ax.text(sp[2][0][0]+Xrd,sp[6][0]/2,t1,horizontalalignment='right',\
                verticalalignment='center',rotation='vertical',fontproperties=font,fontsize=text)
            ax.text(h1_xy[0]+Xrd,(h1_xy[1]-sp[2][1][1])/2+sp[2][1][1],str(round(h1,2))+" м",\
                horizontalalignment='left',verticalalignment='center',rotation='vertical',
                fontproperties=font,fontsize=text)
        if t2!=None:
            ax.text(sp[2][2][0]+Xrd,sp[6][1]/2,t2,horizontalalignment='right',\
                verticalalignment='center',rotation='vertical',fontproperties=font,fontsize=text)
            ax.text(h2_xy[0]+Xrd,(h2_xy[1]-sp[2][3][1])/2+sp[2][3][1],str(round(h2,2))+" м",\
                horizontalalignment='left',verticalalignment='center',rotation='vertical',\
                    fontproperties=font,fontsize=text)
            ax.set_title(title,fontproperties=font,fontsize=tit)

            

        ax.grid(True) # Включаем сетку
        ax.set_xlabel('км',fontproperties=font,fontsize=label_xy) # Подпись оси х
        ax.set_ylabel('м',fontproperties=font,fontsize=label_xy) # Подпись оси у

        for label in (ax.get_xticklabels() + ax.get_yticklabels()):
            label.set_fontproperties(font)
            label.set_fontsize(tiks) # Size here overrides font_prop


        if t2!=None:
            
            xmin,xmax = ax.get_xlim()
            ymin,ymax = ax.get_ylim()
            ax.set_ylim([0, ymax])

            if self.print_param[6] == 1:
                w = self.print_param[7][0]/25.4
                h = self.print_param[7][1]/25.4

                px_text = (self.print_param[2]*self.print_param[4]/100*1.5)
                pxy = h*self.print_param[4]*0.85
                t_y = int(pxy/px_text)
                y_stp = ymax/t_y
                dy=(int(y_stp//5)+1)*5
                ax.set_yticks(np.arange(0,ymax,dy))

                pxx = w*self.print_param[4]*0.85
                t_x = int(pxx/px_text) #*3.3
                x_stp = xmax/t_x
                dx = (int(x_stp//0.25)+1)*0.25
                ax.set_xticks(np.arange(0,xmax,dx))
                ax.tick_params(axis='x', labelrotation=90)
                ax.xaxis.set_major_formatter(FormatStrFormatter('%0.2f'))

                mn = (int((pxx/t_y)//px_text)+0)*dx
                ax.set_xlim([0-mn, sp[2][4][np.shape(sp[2][4])[0]-1]+Xrd+mn])

                fig.legend(handles=(line1, line2, line3,line4,line5),\
                        labels=('Вышки', 'Высота\nн.у.м.', 'Rземли','+'+str(zapas)+' м','Сигнал'),
                        loc='right',prop=font)#

            elif self.print_param[6] == 2:
                y_stp = int(ymax/self.print_param[7][1])
                x_stp = int(xmax/(self.print_param[7][0]/1000))

                w = x_stp/2.54/0.85
                h = y_stp/2.54/0.85

                ax.set_yticks(np.arange(0,ymax,self.print_param[7][1]))
                ax.set_xticks(np.arange(0,xmax,self.print_param[7][0]/1000))
                ax.tick_params(axis='x', labelrotation=90)
                ax.xaxis.set_major_formatter(FormatStrFormatter('%0.2f'))

                mn = (self.print_param[2]*self.print_param[4]/100*1.5)/self.print_param[4]*2.54*self.print_param[7][0]/1000
                
                ax.axis([0-mn*2,sp[2][4][np.shape(sp[2][4])[0]-1]+Xrd+mn*2,0,ymax],'equal')

                ax.legend(handles=(line1, line2, line3,line4,line5),\
                        labels=('Вышки', 'Высота\nн.у.м.', 'Rземли','+'+str(zapas)+' м','Сигнал'),
                        loc='best',prop=font)#
                

            if self.print_param[0]=='Отдельные':
                path_name = self.print_param[8]+"/"+title+" ("+t1+' - '+t2+")"+"."+self.print_param[5]
            else:
                path_name = self.print_param[8]+"/"+title+"."+self.print_param[5]
            fig.set_size_inches(w, h,forward=True) # Изменяем размер сохраняемого графика
            try:
                fig.savefig(path_name, format=self.print_param[5], dpi=self.print_param[4])
            except Exception as ex:
                ems = QErrorMessage(self)
                ems.setWindowTitle('Возникла ошибка')
                ems.showMessage('Возможно открыт файл в который осуществлиется попытка сохранения. Закройте файл. ('+str(ex)+')')

        
    def WinPlot(self,mess):
        sp_list, sp_route, d, i = self.potokRasch.OutputDate()
        sp_list, sp_route, i =sp_list[0], sp_route[0], i[0]
        self.Progress.close()
        try:
            if mess!="Good": raise Exception(mess)
            Xrd = 0
            tic = time()
            self.fig.clear()
            #self.fig.add_subplot(111).clear()
            h_m = np.zeros(len(sp_route),dtype=np.float64)
            ln = len(sp_route)
            j=0
            for sp in sp_list:
                j+=1

                h2=main.H_Min(sp[2][6]*1000,sp[2][7]+self.d_zp.value(),sp[3][0](sp_route[j-1][3]),sp[5],self.d_fg.value(),self.d_fgps.value(),"h2",f=sp[3][3])
                h1=main.H_Min(sp[2][6]*1000,sp[2][7]+self.d_zp.value(),sp[3][1](sp_route[j][3]),sp[4],self.d_fg.value(),self.d_fgps.value(),"h1",f=sp[3][2])

                h_m[j-1]=max(h_m[j-1],h1)
                h_m[j]=h2

                self.Table.item(d[sp_route[j-1][0]],9).setText(str(round(h_m[j-1],2)))
                self.Table.item(d[sp_route[j][0]],9).setText(str(round(h_m[j],2)))
                self.Table.item(d[sp_route[j-1][0]],8).setText(str(round(sp[6][0],2)))
                self.Table.item(d[sp_route[j][0]],8).setText(str(round(sp[6][1],2)))

                Xrd += sp[0]/2
                if j == ln-1:
                    self.Relef(sp,Xrd,sp_route[j-1][3],sp_route[j][3],self.d_zp.value(),\
                        t1=sp_route[j-1][1],t2=sp_route[j][1],title=self.t_v[i][1])
                else:
                    self.Relef(sp,Xrd,sp_route[j-1][3],sp_route[j][3],self.d_zp.value(),\
                        t1=sp_route[j-1][1])
                Xrd += sp[0]/2

            tuc = time()
            print(tuc-tic)
            print("Завершаем расчёт")
            print("-"*5)
            self.global_bloc=False
        except Exception as ex:
            self.global_bloc=False
            if str(ex) == "Elevation not found":
                ems = QErrorMessage(self)
                ems.setWindowTitle('Возникла ошибка')
                ems.showMessage('Не данных о высоте на маршруте выбранного пути. Добавте сотвестующие маршруту файлы в папку hgt.')
            else:
                print(ex)

    def SPlot(self):
        """ Диологовое окно сохранения графиков """
        self.ZapisV()
        j=-1
        for i in range(len(self.t_v)):
            if self.t_v[i][0]:
                j=i
                break

        D = SavePlots(j,self.tv_ind+1,lambda x:self.Make_Plot(x),parent=self)
        D.show()

    def Make_Plot(self,param):
        if param[6] ==3:
            fname = QFileDialog.getSaveFileName(self, 'Сохранить файл', self.path_home,'*.xlsx;;*.xls')[0]
        else:
            fname = QFileDialog.getExistingDirectory(self, 'Сохранить файлы', self.path_home)
        if fname =="": return
        sp_route = []
        self.print_param = param+[fname]
        for i in range(len(param[3])):
            sp_route.append(self.Calculate(i=param[3][i],call=True,prnt=True))


        d = {v:k for k,v in self.id_t.items()}

        self.potokRasch.InputDate(sp_route,d,param[3],True)
        print("Запускаем расчёт")
        if not self.potokRasch.isRunning():
            self.makeplot = False
            self.potokRasch.start()
            self.Progress.show()

    def PointElevSaveExcel(self,wb,title,t1,t2,sp,ws,newpage=False):
        if ws==None:
            ws = []
        
        if newpage:
            if len(ws) == 0:
                ws.append(wb.active)
            else:
                ws.append(wb.create_sheet())
            try:
                ws[len(ws)-1].title = title
            except Exception:
                ws[len(ws)-1].title = str(len(ws))

        lw = len(ws)-1
        ws[lw].append([t1,t2,"","","","","","","",""])
        ws[lw].append(['Длинна, км','Высота, м','Широта, \u00B0','Долгота, \u00B0',\
            'Широта, \u00B0','Минуты',"Секунды",'Долгота, \u00B0',"Минуты","Секунды"])

        ln = np.shape(sp[7])[0]
        for j in range(ln):
            lat= main.DD_in_DMS(sp[7][j][0])
            lon= main.DD_in_DMS(sp[7][j][1])
            if j==0:
                ws[lw].append([0,sp[9][j],sp[7][j][0],sp[7][j][1],lat[0],lat[1],lat[2],lon[0],lon[1],lon[2]])
            elif j==ln-1:
                ws[lw].append([sp[1],sp[9][j],sp[7][j][0],sp[7][j][1],lat[0],lat[1],lat[2],lon[0],lon[1],lon[2]])
            else:
                ws[lw].append([sp[8][j-1],sp[9][j],sp[7][j][0],sp[7][j][1],lat[0],lat[1],lat[2],lon[0],lon[1],lon[2]])

        return ws

    def closeEvent(self, event):
        Message = QMessageBox(QMessageBox.Question,  'Выход из программы',
            "Вы дейстивлеьно хотите выйти?", parent=self)
        Message.addButton('Да', QMessageBox.YesRole)
        Message.addButton('Нет', QMessageBox.NoRole)
        Message.addButton('Сохранить', QMessageBox.ActionRole)
        reply = Message.exec()
        if reply == 0:
            qApp.quit()
        elif reply == 1:
            event.ignore()
        elif reply == 2:
            self.SaveFile()
        

class OneRasch(QThread):
    """ Поток для расчёта графиков """
    mysignal = pyqtSignal(str)
    def __init__(self,kesh_calc,step,parent=None):
        QThread.__init__(self, parent)
        self.kesh_calc=kesh_calc
        self.step=step
    def run(self):
        self.rez = []
        for n in range(len(self.sp_route)):
            a_rez=[]
            try:
                print("Выполняем расчёт")
                step = self.step.value()
                ln = len(self.sp_route[n])
                for j in range(1,ln):
                    if (self.sp_route[n][j-1][2],self.sp_route[n][j][2],step) not in self.kesh_calc:
                        sp = main.Grafik(self.sp_route[n][j-1][2],self.sp_route[n][j][2],step=step/1000)
                        self.kesh_calc[(self.sp_route[n][j-1][2],self.sp_route[n][j][2],step)] = sp
                    else:
                        sp = self.kesh_calc[(self.sp_route[n][j-1][2],self.sp_route[n][j][2],step)]
                    a_rez.append(sp)

                
                if not self.trig: self.mysignal.emit('Good')
            except Exception as ex:
                if not self.trig: self.mysignal.emit(str(ex))
            finally:
                self.rez.append(a_rez)
        
        if self.trig: self.mysignal.emit('End')


    def InputDate(self,sp_route,d,i,trig):
        self.sp_route = sp_route
        self.d=d
        self.i=i
        self.trig = trig
    def OutputDate(self):
        return self.rez, self.sp_route,self.d,self.i
        
    def stop( self ):
        print('stop')
        self.terminate()
        self.wait()
        return 'stop'



if __name__=='__main__':
    app=QApplication(sys.argv)
    window = MyWindow()
    window.show()
    sys.exit(app.exec_())